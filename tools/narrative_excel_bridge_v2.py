#!/usr/bin/env python3
"""Narrative Excel ↔ JSON bridge v2: Domain-agnostic, moon-native, roundtrip-safe.

This is the next-generation bridge supporting:
- Full moon runtime schema (characters, channels, cases, conversations, resourceOutcomes, etc.)
- Optional legacy compatibility (v1 911 mode for migration)
- Roundtrip: Excel → JSON → Excel without data loss
- Multi-language and multi-domain support
- Writer-friendly validation and column hints

Usage:
  python tools/narrative_excel_bridge_v2.py export --json data/game-script.json --xlsx data/game-script-editor-v2.xlsx
  python tools/narrative_excel_bridge_v2.py import --xlsx data/game-script-editor-v2.xlsx --json-out data/game-script.json --json-base data/game-script.json
  python tools/narrative_excel_bridge_v2.py template --xlsx data/game-script-editor-v2.xlsx
"""

from __future__ import annotations

import argparse
import copy
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple, Optional
from collections import defaultdict
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

SHEETS_ORDER_V2 = [
    "HOME",
    "CORE_META",
    "VALIDATION",
    "CHARACTERS",
    "CASES",
    "CASE_CHARACTERS",
    "CONVERSATIONS",
    "DIALOGUES",
    "SIGNALS_INFERENCE",
    "INTENTS_ALIASES",
    "CHANNELS",
    "RESOURCES",
    "RESOURCE_CAPACITY",
    "RESOURCE_OUTCOMES",
    "OUTCOME_MESSAGES",
    "ASSETS",
    "ENDINGS",
    "EVENTS_CINEMATICS",
]

CASE_HEADER_ALIASES: Dict[str, List[str]] = {
    "case_id": ["case_id", "id_caso", "id"],
    "title": ["title", "titulo", "nombre_caso", "nombre"],
    "status": ["status", "estado"],
    "category": ["category", "categoria", "tipo_caso", "tipo"],
    "schedule_time": ["schedule_time", "scheduleTime", "hora", "hora_programada", "time", "time_hhmm"],
    "allocation_bucket": ["allocation_bucket", "allocationBucket", "resource_bucket", "bucket", "grupo_recurso"],
    "minimum_clues": ["minimum_clues", "pistas_minimas", "min_clues", "minimum_signals"],
    "severity": ["severity", "severidad", "prioridad"],
    "stress": ["stress", "estres", "intensidad"],
    "opening_line": ["opening_line", "openingLine", "linea_inicial", "intro", "narrative_intro"],
    "optimal_resources": ["optimal_resources", "optimal", "recommended_resources"],
    "narrative_intro": ["narrativeIntro", "narrative_intro", "intro_text"],
}

SIGNAL_HEADER_ALIASES: Dict[str, List[str]] = {
    "signal_key": ["signal_key", "keyword", "clave", "intencion", "intent"],
    "case_id": ["case_id", "id_caso"],
    "reveal_text": ["reveal_text", "response", "mensaje", "texto_revelado"],
    "is_critical": ["is_critical", "is_required", "critica", "critical"],
}

SHEET_NAME_ALIASES: Dict[str, List[str]] = {
    "INTENTS_ALIASES": ["INTENTS_ALIASES", "SEMANTIC_INTENTS"],
    "EVENTS_CINEMATICS": ["EVENTS_CINEMATICS", "CINEMATIC_EVENTS"],
}

CHARACTER_HEADER_ALIASES: Dict[str, List[str]] = {
    "character_id": ["character_id"],
    "name": ["display_name", "name"],
    "role": ["role"],
    "voice": ["voice_style", "voice"],
    "description": ["description"],
}

CHANNEL_HEADER_ALIASES: Dict[str, List[str]] = {
    "channel_id": ["channel_id"],
    "type": ["channel_type", "type"],
    "actor": ["actor_id", "actor"],
    "label": ["ui_label", "label"],
    "color": ["ui_color", "color"],
    "description": ["description"],
}

CONVERSATION_HEADER_ALIASES: Dict[str, List[str]] = {
    "case_id": ["case_id"],
    "channel": ["channel_id", "channel"],
    "actor": ["actor_id", "actor"],
    "time": ["time_hhmm", "time"],
    "type": ["message_type", "type"],
    "content": ["content", "text"],
    "src": ["media_src", "src"],
    "alt": ["media_alt", "alt"],
}

SCHEMA_VERSION = "2.0"
COMPAT_RUNTIME_VERSION = "2.0"
RUNTIME_CORES = [
    "conversation_core",
    "signal_inference_core",
    "domain_adapter",
    "runtime_ui",
]

HEADER_FILL = PatternFill(fill_type="solid", start_color="1F2B3D", end_color="1F2B3D")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
COMMENT_FILL = PatternFill(fill_type="solid", start_color="FFFFCC", end_color="FFFFCC")
COMMENT_FONT = Font(color="666666", size=9, italic=True)
EXAMPLE_FILL = PatternFill(fill_type="solid", start_color="E8F4F8", end_color="E8F4F8")
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ============================================================================
# UTILITIES
# ============================================================================

def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

def csv_join(values: Iterable[Any]) -> str:
    return ", ".join(str(v).strip() for v in values if str(v).strip())

def csv_split(text: Any) -> List[str]:
    if text is None:
        return []
    return [part.strip() for part in str(text).split(",") if part.strip()]

def delimited_split(text: Any, pattern: str = r"[;,]") -> List[str]:
    if text is None:
        return []
    return [part.strip() for part in re.split(pattern, str(text)) if part and str(part).strip()]

def to_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return " | ".join(str(item).strip() for item in value if str(item).strip())
    return str(value)

def as_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or str(value).strip() == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default

def as_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default

def as_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "verdadero", "si", "sí", "yes", "v"}:
        return True
    if text in {"0", "false", "falso", "no", "f"}:
        return False
    return default

def iter_outcomes(resource_outcomes: Dict[str, Any]) -> Iterable[Tuple[str, Dict[str, Any]]]:
    """Yield normalized (outcome_id, outcome_data) from supported structures."""
    if not isinstance(resource_outcomes, dict):
        return
    for key, value in resource_outcomes.items():
        if isinstance(value, list):
            resource_id = key
            for idx, item in enumerate(value, start=1):
                if not isinstance(item, dict):
                    continue
                branch = "partial" if item.get("partial") else ("success" if item.get("success") else "failure")
                case_id = str(item.get("caseId") or item.get("case_id") or "").strip()
                generated_id = str(item.get("id") or "").strip() or f"{resource_id}_{case_id or 'case'}_{branch}_{idx}"
                normalized = dict(item)
                normalized.setdefault("resource", resource_id)
                normalized.setdefault("branch", branch)
                normalized.setdefault("next_event", item.get("nextEvent", ""))
                yield generated_id, normalized
        elif isinstance(value, dict):
            yield key, value

# ============================================================================
# SHEET BUILDERS (v2 SCHEMA)
# ============================================================================

def ensure_sheet(wb: Workbook, name: str, headers: List[str]) -> Worksheet:
    """Create sheet with styled headers."""
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        ws.column_dimensions[get_column_letter(idx)].width = 22
    ws.freeze_panes = "A2"
    return ws

def finalize_sheet(ws: Worksheet) -> None:
    """Apply formatting to data rows."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER_THIN

def add_comment_row(ws: Worksheet, comment_text: str) -> None:
    """Add a comment/instruction row."""
    row_num = ws.max_row + 1
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = comment_text if col == 1 else ""
        cell.fill = COMMENT_FILL
        cell.font = COMMENT_FONT
        cell.alignment = WRAP

def add_example_row(ws: Worksheet, example_data: Dict[str, Any]) -> None:
    """Add an example/template row."""
    headers = [str(c.value) for c in ws[1]]
    row_num = ws.max_row + 1
    for col_idx, header in enumerate(headers, start=1):
        value = example_data.get(header, "")
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = value
        cell.fill = EXAMPLE_FILL
        cell.alignment = WRAP

def apply_editor_guided_validations(wb: Workbook) -> None:
    """Apply data validation rules that guide non-technical Excel authors."""
    max_row = 2000

    cases = wb["CASES"] if "CASES" in wb.sheetnames else None
    if cases is not None:
        dv_status = DataValidation(type="list", formula1='"pending,resolved"', allow_blank=True)
        dv_severity = DataValidation(type="list", formula1='"baja,media,alta,critica"', allow_blank=True)
        dv_minimum = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True)
        dv_stress = DataValidation(type="whole", operator="between", formula1="0", formula2="10", allow_blank=True)
        dv_time = DataValidation(
            type="custom",
            formula1='=AND(LEN(E2)=5,MID(E2,3,1)=":",ISNUMBER(--LEFT(E2,2)),ISNUMBER(--RIGHT(E2,2)),--LEFT(E2,2)<24,--RIGHT(E2,2)<60)',
            allow_blank=True,
        )

        dv_status.error = "Use pending o resolved"
        dv_severity.error = "Use baja, media, alta o critica"
        dv_minimum.error = "minimum_clues debe ser un entero >= 1"
        dv_stress.error = "stress debe estar entre 0 y 10"
        dv_time.error = "schedule_time debe ser HH:MM (24h)"

        for dv in [dv_status, dv_severity, dv_minimum, dv_stress, dv_time]:
            dv.showErrorMessage = True
            cases.add_data_validation(dv)

        dv_status.add(f"C2:C{max_row}")
        dv_severity.add(f"G2:G{max_row}")
        dv_minimum.add(f"F2:F{max_row}")
        dv_stress.add(f"H2:H{max_row}")
        dv_time.add(f"E2:E{max_row}")

    case_characters = wb["CASE_CHARACTERS"] if "CASE_CHARACTERS" in wb.sheetnames else None
    if case_characters is not None:
        dv_primary = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
        dv_primary.error = "is_primary debe ser true o false"
        dv_primary.showErrorMessage = True
        case_characters.add_data_validation(dv_primary)
        dv_primary.add(f"D2:D{max_row}")

    resources = wb["RESOURCES"] if "RESOURCES" in wb.sheetnames else None
    if resources is not None:
        dv_mode = DataValidation(type="list", formula1='"global,bucketed"', allow_blank=True)
        dv_mode.error = "allocation_mode debe ser global o bucketed"
        dv_mode.showErrorMessage = True
        resources.add_data_validation(dv_mode)
        dv_mode.add(f"G2:G{max_row}")

    signals = wb["SIGNALS_INFERENCE"] if "SIGNALS_INFERENCE" in wb.sheetnames else None
    if signals is not None:
        dv_critical = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
        dv_critical.error = "is_critical debe ser true o false"
        dv_critical.showErrorMessage = True
        signals.add_data_validation(dv_critical)
        dv_critical.add(f"D2:D{max_row}")

# ============================================================================
# SHEET: HOME (Editorial Navigation & Authoring Guide)
# ============================================================================

def write_home(ws: Worksheet) -> None:
    """Create HOME as editorial UX hub: navigation, filling guide, and practical examples."""
    current_row = 1

    def add_section(title: str) -> None:
        nonlocal current_row
        ws.cell(row=current_row, column=1).value = title
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=current_row, column=1).fill = HEADER_FILL
        ws.merge_cells(f"A{current_row}:F{current_row}")
        current_row += 1

    # Title
    ws.cell(row=current_row, column=1).value = "HOME · Corazón editorial del workbook"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(fill_type="solid", start_color="1F2B3D", end_color="1F2B3D")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1

    ws.cell(row=current_row, column=1).value = "Navega, aprende y authora desde aquí. Cada bloque explica qué llenar y por qué impacta el juego."
    ws.cell(row=current_row, column=1).font = Font(italic=True, size=10, color="666666")
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 2

    # Navigation section with clickable links
    add_section("NAVEGACIÓN RÁPIDA (clic para abrir pestaña)")
    ws.cell(row=current_row, column=1).value = "Pestaña"
    ws.cell(row=current_row, column=2).value = "Para qué sirve"
    ws.cell(row=current_row, column=3).value = "Qué completar primero"
    ws.cell(row=current_row, column=4).value = "Impacto en runtime"
    ws.cell(row=current_row, column=5).value = "Ir"
    ws.cell(row=current_row, column=6).value = "Prioridad"
    for col in range(1, 7):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = EXAMPLE_FILL
        cell.font = Font(bold=True, size=10)
    current_row += 1

    nav_rows = [
        ("CORE_META", "Config global del juego", "domain, title, strict_authoring, locales", "Define etiquetas, idioma y modo estricto", "1"),
        ("CHARACTERS", "Personajes y voces", "character_id, display_name, role", "Quién habla y cómo se ve en UI", "2"),
        ("CHANNELS", "Canales de comunicación", "channel_id, channel_type, actor_id", "Dónde aparecen mensajes (chat, newsletter, etc.)", "3"),
        ("CASES", "Casos narrativos", "case_id, title, schedule_time, severity", "Secuencia narrativa jugable", "4"),
        ("CASE_CHARACTERS", "Vínculo explícito caso↔personaje", "case_id + character_id", "Sin vínculos no hay coherencia narrativa", "5"),
        ("CONVERSATIONS", "Mensajes cronológicos", "actor_id, time_hhmm, content", "Flujo de chat en escena", "6"),
        ("DIALOGUES", "Respuestas a intención/pista", "intent, keyword, response, clue", "Qué obtiene el jugador al preguntar", "7"),
        ("SIGNALS_INFERENCE", "Mapa señal→revelación", "signal_key, reveal_text, is_critical", "Progresión de pistas", "8"),
        ("RESOURCES", "Acciones/intervenciones", "resource_id, allocation_mode", "Botones de intervención en runtime", "9"),
        ("RESOURCE_CAPACITY", "Capacidad por bucket", "resource_id, bucket_key, available", "Cuántas veces puedes usar un recurso", "10"),
        ("RESOURCE_OUTCOMES", "Resultado de cada recurso", "branch, confidence_band, score_delta", "Consecuencias jugables", "11"),
        ("OUTCOME_MESSAGES", "Mensajes post-acción", "delay_ms, actor_id, content", "Feedback narrativo tras decidir", "12"),
        ("ENDINGS", "Finales", "condition_expr, ending_type", "Cierre de partida", "13"),
        ("VALIDATION", "Chequeos editoriales", "Revisar errores de referencias", "Evita importaciones rotas", "14"),
    ]

    for sheet_name, purpose, first_fill, runtime_impact, priority in nav_rows:
        ws.cell(row=current_row, column=1).value = sheet_name
        ws.cell(row=current_row, column=2).value = purpose
        ws.cell(row=current_row, column=3).value = first_fill
        ws.cell(row=current_row, column=4).value = runtime_impact
        ws.cell(row=current_row, column=6).value = priority
        link_cell = ws.cell(row=current_row, column=5)
        link_cell.value = f"Abrir {sheet_name}"
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color="0563C1", underline="single", bold=True)
        current_row += 1

    current_row += 1
    add_section("FLUJO DE LLENADO RECOMENDADO")
    flow_steps = [
        "1) CORE_META → CHARACTERS → CHANNELS",
        "2) CASES → CASE_CHARACTERS → CONVERSATIONS",
        "3) DIALOGUES + SIGNALS_INFERENCE",
        "4) RESOURCES + RESOURCE_CAPACITY + RESOURCE_OUTCOMES",
        "5) OUTCOME_MESSAGES + ENDINGS",
        "6) VALIDATION y luego import al JSON",
    ]
    for step in flow_steps:
        ws.cell(row=current_row, column=1).value = step
        ws.merge_cells(f"A{current_row}:F{current_row}")
        current_row += 1

    current_row += 1
    add_section("CASOS DE USO PRÁCTICOS (sólidos y compuestos)")

    # Use case 1
    ws.cell(row=current_row, column=1).value = "Caso 1 · Hacer que un personaje hable por cierto canal"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1
    ws.cell(row=current_row, column=1).value = "CHARACTERS"
    ws.cell(row=current_row, column=2).value = "character_id=caller_ana, display_name=Ana"
    ws.cell(row=current_row, column=3).value = "CHANNELS"
    ws.cell(row=current_row, column=4).value = "channel_id=chat_ana, channel_type=chat, actor_id=caller_ana"
    ws.cell(row=current_row, column=5).value = "CONVERSATIONS"
    ws.cell(row=current_row, column=6).value = "actor_id=caller_ana, channel_id=chat_ana, content='Te cuento algo...'"
    current_row += 1

    # Use case 2
    ws.cell(row=current_row, column=1).value = "Caso 2 · Configurar respuestas casuales"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1
    ws.cell(row=current_row, column=1).value = "CORE_META"
    ws.cell(row=current_row, column=2).value = "no_match_responses='mmm, no te sigo, dame un dato; ok, necesito algo más concreto'"
    ws.cell(row=current_row, column=3).value = "DIALOGUES"
    ws.cell(row=current_row, column=4).value = "intent=casual_check, keyword=hola, response='Hola, ¿qué necesitas confirmar?'"
    ws.merge_cells(f"E{current_row}:F{current_row}")
    current_row += 1

    # Use case 3
    ws.cell(row=current_row, column=1).value = "Caso 3 · Revelar una pista en un diálogo"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1
    ws.cell(row=current_row, column=1).value = "DIALOGUES"
    ws.cell(row=current_row, column=2).value = "case_id=reveal_truth_01, keyword=nervioso, clue=temblor_voz"
    ws.cell(row=current_row, column=3).value = "SIGNALS_INFERENCE"
    ws.cell(row=current_row, column=4).value = "signal_key=temblor_voz, reveal_text='Ana evita una parte clave', is_critical=true"
    ws.merge_cells(f"E{current_row}:F{current_row}")
    current_row += 1

    # Use case 4
    ws.cell(row=current_row, column=1).value = "Caso 4 · Enviar un WhatsApp o audio programado"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1
    ws.cell(row=current_row, column=1).value = "CHANNELS"
    ws.cell(row=current_row, column=2).value = "channel_id=wa_ana, channel_type=whatsapp, actor_id=caller_ana"
    ws.cell(row=current_row, column=3).value = "CONVERSATIONS"
    ws.cell(row=current_row, column=4).value = "time_hhmm=22:14, message_type=audio, media_src=audio/ana_note.ogg"
    ws.cell(row=current_row, column=5).value = "OUTCOME_MESSAGES"
    ws.cell(row=current_row, column=6).value = "delay_ms=120000 para enviarlo 2 minutos después"
    current_row += 1

    # Use case 5
    ws.cell(row=current_row, column=1).value = "Caso 5 · Dos personajes envían la misma imagen con desfase"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1
    ws.cell(row=current_row, column=1).value = "CONVERSATIONS fila A"
    ws.cell(row=current_row, column=2).value = "actor_id=caller_ana, message_type=image, media_src=images/vignettes/clue_01.jpg, time_hhmm=22:18"
    ws.cell(row=current_row, column=3).value = "CONVERSATIONS fila B"
    ws.cell(row=current_row, column=4).value = "actor_id=caller_diego, message_type=image, media_src=images/vignettes/clue_01.jpg, time_hhmm=22:21"
    ws.cell(row=current_row, column=5).value = "Resultado"
    ws.cell(row=current_row, column=6).value = "Misma imagen, desfase de 3 minutos para tensión narrativa"
    current_row += 1

    current_row += 1
    add_section("PRINCIPIOS CLAVE")
    principles = [
        "No heurísticas ocultas: todo vínculo y regla debe quedar escrito en hojas.",
        "CASE_CHARACTERS es obligatorio para coherencia de actores por caso.",
        "allocation_bucket y RESOURCE_CAPACITY permiten escalar a múltiples clientes.",
        "Excel es fuente de verdad; el JSON se regenera desde este workbook.",
        "Si algo no está en Excel, no existe editorialmente en runtime.",
    ]
    for principle in principles:
        ws.cell(row=current_row, column=1).value = f"• {principle}"
        ws.merge_cells(f"A{current_row}:F{current_row}")
        current_row += 1

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 44
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 42

    finalize_sheet(ws)

# ============================================================================
# SHEET: CORE_META
# ============================================================================

def write_core_meta(ws: Worksheet) -> None:
    """Configuration metadata for the entire game."""
    add_comment_row(ws, "KEY=VALUE pairs for game metadata. Editable.")
    add_example_row(ws, {
        "key": "game_domain",
        "value": "mimi",
        "description": "Name of game domain (e.g., mimi, 911, sales, support).",
    })
    
    meta_rows = [
        ("game_domain", "mimi", "Name of this game domain (no spaces, lowercase)"),
        ("game_title", "MIMI: Misión Mentira", "Display title for players"),
        ("game_version", "0.1.0", "Semantic version (MAJOR.MINOR.PATCH)"),
        ("strict_authoring", "true", "true = disable hidden heuristic/autogenerated runtime behavior"),
        ("player_label", "Operador", "Label shown for player utterances"),
        ("turn_start_time", "22:00", "Shift start HH:MM for runtime clock"),
        ("turn_end_time", "01:30", "Shift end HH:MM for runtime clock"),
        ("clock_tick_interval_ms", "7000", "Real milliseconds per game-minute tick window"),
        ("clock_advance_slow_minutes", "4", "Game minutes advanced per tick in slow pace"),
        ("clock_advance_normal_minutes", "6", "Game minutes advanced per tick in normal pace"),
        ("clock_advance_rush_minutes", "9", "Game minutes advanced per tick in rush pace"),
        ("no_match_responses", "No te sigo, ¿puedes darme un dato concreto?", "Comma-separated fallback replies"),
        ("schema_version", SCHEMA_VERSION, "Schema version (read-only reference)"),
        ("runtime_version", COMPAT_RUNTIME_VERSION, "Runtime compatibility version"),
        ("locales", "es, en", "Comma-separated language codes"),
        ("default_locale", "es", "Default language"),
        ("game_loop_duration_minutes", "60", "Estimated playtime in minutes"),
    ]
    
    for key, value, description in meta_rows:
        ws.append([key, value, description])
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: CHARACTERS
# ============================================================================

def write_characters(ws: Worksheet) -> None:
    """Character definitions (callers, UI actors, NPCs)."""
    add_comment_row(ws, "Define all named actors in the game. Each character is unique.")
    add_example_row(ws, {
        "character_id": "mimi_protagonist",
        "display_name": "MIMI",
        "role": "protagonist",
        "voice_style": "neutral, slightly sarcastic",
        "description": "AI assistant helping with truth detection",
    })
    
    example_rows = [
        {
            "character_id": "caller_diego",
            "display_name": "Diego",
            "role": "caller",
            "voice_style": "hesitant, nervous",
            "description": "Teenager with secret about truth claim",
        },
        {
            "character_id": "caller_ana",
            "display_name": "Ana",
            "role": "caller",
            "voice_style": "authoritative, skeptical",
            "description": "Friend who initiated the game",
        },
    ]
    
    for example in example_rows:
        add_example_row(ws, example)
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: CHANNELS
# ============================================================================

def write_channels(ws: Worksheet) -> None:
    """Communication channels (chat, call, broadcast, etc)."""
    add_comment_row(ws, "Define how characters interact. Type determines UI rendering.")
    add_example_row(ws, {
        "channel_id": "chat_diego",
        "channel_type": "direct_message",
        "actor_id": "caller_diego",
        "ui_label": "Diego",
        "ui_color": "#d9879b",
        "description": "Text conversation with Diego",
    })
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: CASES
# ============================================================================

def write_cases(ws: Worksheet) -> None:
    """Narrative units (cases, scenarios, scenes)."""
    add_comment_row(ws, "Define playable cases only. Keep rows short and operational for runtime/UI.")
    add_comment_row(ws, "Recommended: one row per case with clear title, category and schedule_time.")
    add_example_row(ws, {
        "case_id": "reveal_truth_01",
        "title": "¿La verdad de Diego?",
        "status": "pending",
        "category": "investigation",
        "schedule_time": "21:30",
        "allocation_bucket": "chat_priority",
        "minimum_clues": "2",
        "severity": "media",
        "stress": "6",
        "opening_line": "Diego llama con un secreto y duda al hablar.",
    })
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: CASE_CHARACTERS
# ============================================================================

def write_case_characters(ws: Worksheet) -> None:
    """Explicit many-to-many relation between cases and characters."""
    add_comment_row(ws, "Define which characters are linked to each case. This is the canonical relation.")
    add_example_row(ws, {
        "case_id": "reveal_truth_01",
        "character_id": "caller_diego",
        "involvement_role": "caller",
        "is_primary": "true",
    })
    add_example_row(ws, {
        "case_id": "reveal_truth_01",
        "character_id": "mimi_protagonist",
        "involvement_role": "operator",
        "is_primary": "false",
    })
    finalize_sheet(ws)

# ============================================================================
# SHEET: DIALOGUES
# ============================================================================

def write_dialogues(ws: Worksheet) -> None:
    """Dialogue turns tied to case and character with signal/reveal semantics."""
    add_comment_row(ws, "Each row is a dialogue turn authored in Excel (source of truth).")
    add_example_row(ws, {
        "case_id": "reveal_truth_01",
        "character_id": "caller_diego",
        "intent": "ask_truth",
        "keyword": "hesitation",
        "response": "Diego pausa 3 segundos antes de responder.",
        "clue": "hesitation",
        "resource_id": "hint_emotional_read",
    })
    finalize_sheet(ws)

# ============================================================================
# SHEET: CONVERSATIONS
# ============================================================================

def write_conversations(ws: Worksheet) -> None:
    """Scheduled and triggered messages."""
    add_comment_row(ws, "Messages, media, and turn-taking. order_index determines sequence.")
    add_example_row(ws, {
        "conversation_id": "msg_001",
        "case_id": "reveal_truth_01",
        "channel_id": "chat_diego",
        "order_index": "1",
        "actor_id": "caller_diego",
        "time_hhmm": "21:30",
        "message_type": "text",
        "content": "Hola... tengo un secreto.",
        "media_src": "",
        "media_alt": "",
        "visible_if_result": "",
    })
    
    add_comment_row(ws, "Add columns: visible_if_result to control conditional messages.")
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: INTENTS_ALIASES
# ============================================================================

def write_intents_aliases(ws: Worksheet) -> None:
    """Semantic mapping: intents and aliases for natural language matching."""
    add_comment_row(ws, "Canonical intent → aliases. Maps player input to signal keywords.")
    add_example_row(ws, {
        "canonical_intent": "ask_truth",
        "display_label": "Pregunta por la verdad",
        "aliases": "mentira, verdad, engaño, falso, real",
        "description": "Player asks if statement is truth or lie",
    })
    
    add_example_row(ws, {
        "canonical_intent": "ask_feeling",
        "display_label": "Pregunta cómo se siente",
        "aliases": "cómo se siente, estado emocional, bien mal, nervioso",
        "description": "Player asks about character emotional state",
    })
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: SIGNALS_INFERENCE
# ============================================================================

def write_signals_inference(ws: Worksheet) -> None:
    """Signal definitions and inference configuration."""
    add_comment_row(ws, "Define signal keyword -> reveal text per case. Keep only operational fields.")
    add_example_row(ws, {
        "signal_key": "hesitation",
        "case_id": "reveal_truth_01",
        "reveal_text": "Diego pauses for 3 seconds before answering.",
        "is_critical": "true",
    })
    
    add_example_row(ws, {
        "signal_key": "quick_response",
        "case_id": "reveal_truth_01",
        "reveal_text": "Diego responds immediately without thinking.",
        "is_critical": "false",
    })
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: RESOURCES
# ============================================================================

def write_resources(ws: Worksheet) -> None:
    """Game-domain resources (assets, tools, hints, etc)."""
    add_comment_row(ws, "Define resources that can be allocated/used by player or system.")
    add_comment_row(ws, "Use allocation_mode = global or bucketed. Bucket stock is authored in RESOURCE_CAPACITY.")
    add_example_row(ws, {
        "resource_id": "hint_emotional_read",
        "label": "Emotional Signal Reader",
        "description": "Tool to detect emotional shifts in voice",
        "category": "tool",
        "total_available": "3",
        "cost": "1",
        "allocation_mode": "bucketed",
    })
    
    add_example_row(ws, {
        "resource_id": "hint_body_language",
        "label": "Body Language Monitor",
        "description": "Camera view of caller's body language",
        "category": "sensor",
        "total_available": "1",
        "cost": "1",
        "allocation_mode": "global",
    })
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: RESOURCE_CAPACITY
# ============================================================================

def write_resource_capacity(ws: Worksheet) -> None:
    """Generic capacity buckets for resources, independent of any 911 zone model."""
    add_comment_row(ws, "Optional. One row per resource bucket/capacity pool. Leave empty for global-only resources.")
    add_example_row(ws, {
        "resource_id": "hint_emotional_read",
        "bucket_key": "chat_priority",
        "available": "2",
        "travel_ms": "0",
        "description": "High-priority private chat capacity",
    })
    add_example_row(ws, {
        "resource_id": "hint_emotional_read",
        "bucket_key": "general_queue",
        "available": "1",
        "travel_ms": "0",
        "description": "Shared queue capacity",
    })
    finalize_sheet(ws)

# ============================================================================
# SHEET: RESOURCE_OUTCOMES
# ============================================================================

def write_resource_outcomes(ws: Worksheet) -> None:
    """What happens when resources are used/allocated."""
    add_comment_row(ws, "Outcomes of using resources: branches (success/partial/failure) and rewards.")
    add_example_row(ws, {
        "outcome_id": "outcome_hint_emotional_01",
        "case_id": "reveal_truth_01",
        "resource_id": "hint_emotional_read",
        "branch": "success",
        "confidence_band": "high",
        "unlocks_signals": "hesitation, vocal_nervousness",
        "score_delta": "5",
        "calm_delta": "0",
        "next_event": "player_gains_insight",
        "keep_case_open": "true",
    })
    
    add_example_row(ws, {
        "outcome_id": "outcome_hint_emotional_02",
        "case_id": "reveal_truth_01",
        "resource_id": "hint_emotional_read",
        "branch": "failure",
        "confidence_band": "low",
        "unlocks_signals": "",
        "score_delta": "-2",
        "calm_delta": "-1",
        "next_event": "",
        "keep_case_open": "false",
    })
    
    add_comment_row(ws, "branch: success, partial, failure. confidence_band: low/medium/high.")
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: OUTCOME_MESSAGES
# ============================================================================

def write_outcome_messages(ws: Worksheet) -> None:
    """Messages that result from outcomes."""
    add_comment_row(ws, "Messages triggered by outcomes: delivery, delay, hint/consequence routing.")
    add_example_row(ws, {
        "message_id": "msg_outcome_001",
        "outcome_id": "outcome_hint_emotional_01",
        "channel_id": "chat_diego",
        "actor_id": "caller_diego",
        "delay_ms": "2000",
        "message_type": "text",
        "content": "Notaste que dudé? Bueno...",
        "hint_key": "hesitation_reveals_truth",
        "next_case": "clarify_secret",
    })
    
    add_comment_row(ws, "delay_ms: milliseconds before message appears. hint_key: optional clue label.")
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: EVENTS_CINEMATICS
# ============================================================================

def write_events_cinematics(ws: Worksheet) -> None:
    """Cinematic events, state changes, and narrative progression."""
    add_comment_row(ws, "Non-interactive narrative events: camera changes, music, visual effects, etc.")
    add_example_row(ws, {
        "event_id": "cinematic_diego_revealed",
        "case_id": "reveal_truth_01",
        "trigger_condition": "all_signals_found",
        "event_type": "cinematic",
        "title": "Diego's Truth",
        "body": "After your questions, Diego finally admits the truth...",
        "priority": "1",
        "media_src": "cinematic/diego_truth.mp4",
    })
    
    add_comment_row(ws, "trigger_condition: on_case_start, on_signal_found, all_signals_found, on_outcome_branch, etc.")
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: ENDINGS
# ============================================================================

def write_endings(ws: Worksheet) -> None:
    """Game endings and results."""
    add_comment_row(ws, "Define possible endings based on player outcome and choices.")
    add_example_row(ws, {
        "ending_id": "ending_truth_revealed",
        "condition_expr": "case_reveal_truth_01:resuelto AND signals_found >= 2",
        "ending_type": "victory",
        "title": "You Discovered the Truth",
        "body": "Diego's secret is out. The truth was... [reveal details]",
        "score_delta": "50",
        "next_game": "",
    })
    
    add_example_row(ws, {
        "ending_id": "ending_lie_believed",
        "condition_expr": "case_reveal_truth_01:incomplete",
        "ending_type": "partial",
        "title": "The Lie Stands",
        "body": "You didn't ask the right questions. Diego's story remains unchanged.",
        "score_delta": "0",
        "next_game": "",
    })
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: ASSETS
# ============================================================================

def write_assets(ws: Worksheet) -> None:
    """Media assets: images, audio, video."""
    add_comment_row(ws, "Catalog of all media used. Supports fallback chains.")
    add_example_row(ws, {
        "asset_id": "music_tension",
        "asset_type": "music",
        "path": "assets/music/tension-loop.mp3",
        "license": "CC-BY-4.0",
        "fallback_asset_id": "music_tension_alt",
        "description": "Background music during investigation",
    })
    
    add_example_row(ws, {
        "asset_id": "sfx_notification",
        "asset_type": "sfx",
        "path": "assets/sfx/message-alert.mp3",
        "license": "proprietary",
        "fallback_asset_id": "",
        "description": "Notification sound for new message",
    })
    
    finalize_sheet(ws)

# ============================================================================
# SHEET: VALIDATION
# ============================================================================

def write_validation(ws: Worksheet) -> None:
    """Validation rules and error checking."""
    add_comment_row(ws, "This sheet explains what the importer validates. It is documentation for authors.")
    add_comment_row(ws, "If import reports warnings/errors, fix the referenced sheet and ID relation first.")

    validation_rules = [
        ("Reference", "CONVERSATIONS", "channel_id must exist in CHANNELS", "Channel not found"),
        ("Reference", "CONVERSATIONS", "case_id must exist in CASES", "Case not found"),
        ("Reference", "CASE_CHARACTERS", "case_id must exist in CASES", "Case not found"),
        ("Reference", "CASE_CHARACTERS", "character_id must exist in CHARACTERS", "Character not found"),
        ("Authoring", "CASE_CHARACTERS", "Do not rely on conversations to imply relationships", "Case-character link must be explicit"),
        ("Reference", "DIALOGUES", "(case_id, character_id) must exist in CASE_CHARACTERS", "Dialogue character is not linked to case"),
        ("Authoring", "CASES", "status in {pending, resolved}", "Invalid status value"),
        ("Authoring", "CASES", "schedule_time should be HH:MM", "Invalid schedule format"),
        ("Reference", "SIGNALS_INFERENCE", "case_id must exist in CASES", "Case not found"),
        ("Authoring", "SIGNALS_INFERENCE", "signal_key + reveal_text must be non-empty", "Incomplete signal row"),
        ("Reference", "RESOURCE_CAPACITY", "resource_id must exist in RESOURCES", "Resource not found"),
        ("Enum", "RESOURCE_OUTCOMES", "branch in {success, partial, failure}", "Branch must be success/partial/failure"),
        ("Enum", "EVENTS_CINEMATICS", "event_type in {cinematic, state_change, audio, visual}", "Invalid event type"),
        ("Unique", "CHARACTERS", "character_id is unique across sheet", "Duplicate character_id"),
        ("Unique", "CASES", "case_id is unique across sheet", "Duplicate case_id"),
        ("Range", "RESOURCE_OUTCOMES", "score_delta in [-100, 100]", "Score delta out of range"),
        ("Reference", "OUTCOME_MESSAGES", "outcome_id must exist in RESOURCE_OUTCOMES", "Outcome not found"),
    ]
    
    for rule_type, scope, check, error in validation_rules:
        ws.append([rule_type, scope, check, error])
    
    finalize_sheet(ws)

# ============================================================================
# EXPORT TO WORKBOOK
# ============================================================================

def export_to_workbook_v2(data: Dict[str, Any], xlsx_path: Path) -> None:
    """Export game JSON to v2 Excel template."""
    wb = Workbook()
    wb.remove(wb.active)

    # HOME sheet (editorial navigation)
    ws = wb.create_sheet(title="HOME", index=0)
    write_home(ws)
    
    # CORE_META
    ws = ensure_sheet(wb, "CORE_META", ["key", "value", "description"])
    semantic = (data.get("settings", {}) or {}).get("semantic", {}) or {}
    interaction = (data.get("settings", {}) or {}).get("interaction", {}) or {}
    ui_settings = (data.get("settings", {}) or {}).get("ui", {}) or {}
    for key, value, desc in [
        ("game_domain", data.get("game_domain", ""), ""),
        ("game_title", data.get("game_title", ""), ""),
        ("game_version", data.get("version", ""), ""),
        ("strict_authoring", str(as_bool(interaction.get("strictAuthoring", False))).lower(), "true = no hidden heuristics/autogen"),
        ("player_label", ui_settings.get("playerLabel", "Operador"), "Label shown for player utterances"),
        ("turn_start_time", interaction.get("turnStartTime", "22:00"), "Shift start HH:MM for runtime clock"),
        ("turn_end_time", interaction.get("turnEndTime", "01:30"), "Shift end HH:MM for runtime clock"),
        ("clock_tick_interval_ms", interaction.get("clockTickIntervalMs", 7000), "Real milliseconds per game-minute tick window"),
        ("clock_advance_slow_minutes", interaction.get("clockAdvanceSlowMinutes", 4), "Game minutes advanced per tick in slow pace"),
        ("clock_advance_normal_minutes", interaction.get("clockAdvanceNormalMinutes", 6), "Game minutes advanced per tick in normal pace"),
        ("clock_advance_rush_minutes", interaction.get("clockAdvanceRushMinutes", 9), "Game minutes advanced per tick in rush pace"),
        ("no_match_responses", csv_join(semantic.get("noMatchResponses", [])), "Comma-separated fallback replies"),
    ]:
        ws.append([key, value, desc])
    finalize_sheet(ws)
    
    # CHARACTERS
    ws = ensure_sheet(wb, "CHARACTERS", ["character_id", "display_name", "role", "voice_style", "description"])
    for char_id, char_data in (data.get("characters", {}) or {}).items():
        ws.append([
            char_id,
            char_data.get("name", ""),
            char_data.get("role", ""),
            char_data.get("voice", ""),
            char_data.get("description", ""),
        ])
    finalize_sheet(ws)
    
    # CHANNELS
    ws = ensure_sheet(wb, "CHANNELS", ["channel_id", "channel_type", "actor_id", "ui_label", "ui_color", "description"])
    for channel_id, channel_data in (data.get("channels", []) or []):
        if isinstance(channel_data, dict):
            ws.append([
                channel_id,
                channel_data.get("type", ""),
                channel_data.get("actor", ""),
                channel_data.get("label", ""),
                channel_data.get("color", ""),
                channel_data.get("description", ""),
            ])
    finalize_sheet(ws)
    
    # CASES
    ws = ensure_sheet(wb, "CASES", [
        "case_id", "title", "status", "category", "schedule_time", "allocation_bucket",
        "minimum_clues", "severity", "stress", "opening_line"
    ])
    for case_data in (data.get("cases", []) or []):
        if isinstance(case_data, dict):
            ws.append([
                case_data.get("id", ""),
                case_data.get("title", ""),
                case_data.get("status", "pending"),
                case_data.get("category", ""),
                case_data.get("scheduleTime", ""),
                case_data.get("allocationBucket", case_data.get("resourceBucket", "")),
                case_data.get("minimumClues", len(case_data.get("requiredSignals", []) or [])),
                case_data.get("severity", "media"),
                case_data.get("stress", ""),
                to_cell_text(case_data.get("openingLine") or case_data.get("narrativeIntro") or ""),
            ])
    finalize_sheet(ws)

    # CASE_CHARACTERS
    ws = ensure_sheet(wb, "CASE_CHARACTERS", ["case_id", "character_id", "involvement_role", "is_primary"])
    for case_data in (data.get("cases", []) or []):
        case_id = case_data.get("id", "")
        case_chars = case_data.get("caseCharacters", []) or []
        if case_chars:
            for rel in case_chars:
                if not isinstance(rel, dict):
                    continue
                character_id = str(rel.get("character_id") or rel.get("characterId") or rel.get("id") or "").strip()
                if not character_id:
                    continue
                ws.append([
                    case_id,
                    character_id,
                    rel.get("involvement_role") or rel.get("role") or "participant",
                    str(as_bool(rel.get("is_primary"), False)).lower(),
                ])
            continue
        explicit_ids = case_data.get("characterIds", []) or []
        if explicit_ids:
            for index, character_id in enumerate(explicit_ids):
                ws.append([case_id, character_id, "participant", str(index == 0).lower()])
            continue
    finalize_sheet(ws)

    # DIALOGUES
    ws = ensure_sheet(wb, "DIALOGUES", ["case_id", "character_id", "intent", "keyword", "response", "clue", "resource_id"])
    for case_data in (data.get("cases", []) or []):
        case_id = case_data.get("id", "")
        authored_dialogues = case_data.get("dialogues", []) or []
        if authored_dialogues:
            for row in authored_dialogues:
                if not isinstance(row, dict):
                    continue
                ws.append([
                    case_id,
                    row.get("character_id") or row.get("characterId") or row.get("actor") or "",
                    row.get("intent", ""),
                    row.get("keyword", ""),
                    row.get("response", ""),
                    row.get("clue", ""),
                    row.get("resource_id") or row.get("resourceId") or "",
                ])
            continue
        reveal = case_data.get("signalResponse", {}) or {}
        fallback_character = ""
        case_character_ids = case_data.get("characterIds", []) or []
        if case_character_ids:
            fallback_character = str(case_character_ids[0])
        for keyword, response in reveal.items():
            ws.append([case_id, fallback_character, keyword, keyword, response, keyword, ""])
    finalize_sheet(ws)
    
    # CONVERSATIONS
    ws = ensure_sheet(wb, "CONVERSATIONS", [
        "conversation_id", "case_id", "channel_id", "order_index", "actor_id",
        "time_hhmm", "message_type", "content", "media_src", "media_alt", "visible_if_result"
    ])
    for case_data in (data.get("cases", []) or []):
        for conv_idx, conv in enumerate((case_data.get("conversations", []) or []), start=1):
            if isinstance(conv, dict):
                ws.append([
                    f"{case_data.get('id')}:conv:{conv_idx}",
                    case_data.get("id", ""),
                    conv.get("channel", ""),
                    conv_idx,
                    conv.get("actor", ""),
                    conv.get("time", ""),
                    conv.get("type", "text"),
                    conv.get("content", ""),
                    conv.get("src", ""),
                    conv.get("alt", ""),
                    "",
                ])
    finalize_sheet(ws)
    
    # INTENTS_ALIASES
    ws = ensure_sheet(wb, "INTENTS_ALIASES", ["canonical_intent", "display_label", "aliases", "description"])
    for intent, aliases in ((data.get("settings", {}) or {}).get("semantic", {}).get("keywordAliases", {}) or {}).items():
        alias_list = ", ".join(aliases if isinstance(aliases, list) else [str(aliases)])
        ws.append([intent, intent, alias_list, ""])
    finalize_sheet(ws)
    
    # SIGNALS_INFERENCE
    ws = ensure_sheet(wb, "SIGNALS_INFERENCE", [
        "signal_key", "case_id", "reveal_text", "is_critical"
    ])
    for case_data in (data.get("cases", []) or []):
        for signal_key, reveal_text in ((case_data.get("signalResponse", {}) or {}).items()):
            critical = signal_key in (case_data.get("requiredSignals", []) or [])
            ws.append([
                signal_key,
                case_data.get("id", ""),
                reveal_text,
                str(critical).lower(),
            ])
    finalize_sheet(ws)
    
    # RESOURCES
    ws = ensure_sheet(wb, "RESOURCES", [
        "resource_id", "label", "description", "category", "total_available", "cost", "allocation_mode"
    ])
    for res_id, res_data in (data.get("resources", {}) or {}).items():
        ws.append([
            res_id,
            res_data.get("label", ""),
            res_data.get("description", ""),
            res_data.get("category", "tool"),
            res_data.get("total", 0),
            res_data.get("cost", 1),
            res_data.get("allocationMode", "bucketed" if (res_data.get("capacities") or res_data.get("zones")) else "global"),
        ])
    finalize_sheet(ws)

    ws = ensure_sheet(wb, "RESOURCE_CAPACITY", ["resource_id", "bucket_key", "available", "travel_ms", "description"])
    for res_id, res_data in (data.get("resources", {}) or {}).items():
        capacities = res_data.get("capacities") or res_data.get("zones") or {}
        for bucket_key, amount in (capacities or {}).items():
            ws.append([res_id, bucket_key, amount, "0", ""])
    finalize_sheet(ws)
    
    # RESOURCE_OUTCOMES
    ws = ensure_sheet(wb, "RESOURCE_OUTCOMES", [
        "outcome_id", "case_id", "resource_id", "branch", "confidence_band",
        "unlocks_signals", "score_delta", "calm_delta", "next_event", "keep_case_open"
    ])
    for outcome_id, outcome in iter_outcomes((data.get("resourceOutcomes", {}) or {})):
        resource_id = str(outcome.get("resource") or outcome.get("resource_id") or "").strip()
        case_id = str(outcome.get("case_id") or outcome.get("caseId") or "").strip()
        branch = str(outcome.get("branch") or "").strip().lower()
        if not branch:
            branch = "partial" if as_bool(outcome.get("partial"), False) else ("success" if as_bool(outcome.get("success"), False) else "failure")
        ws.append([
            outcome_id,
            case_id,
            resource_id,
            branch,
            outcome.get("confidence_band", ""),
            csv_join(outcome.get("unlocks") or outcome.get("cluesUnlocked") or []),
            outcome.get("score_delta") if outcome.get("score_delta") is not None else outcome.get("scoreDelta", ""),
            outcome.get("calm_delta") if outcome.get("calm_delta") is not None else outcome.get("calmDelta", ""),
            outcome.get("next_event") if outcome.get("next_event") is not None else outcome.get("nextEvent", ""),
            str(as_bool(outcome.get("keep_case_open") if outcome.get("keep_case_open") is not None else outcome.get("keepCaseOpen"), False)).lower(),
        ])
    finalize_sheet(ws)
    
    # OUTCOME_MESSAGES
    ws = ensure_sheet(wb, "OUTCOME_MESSAGES", [
        "message_id", "outcome_id", "channel_id", "actor_id", "delay_ms",
        "message_type", "content", "hint_key", "next_case"
    ])
    msg_index = 1
    for outcome_id, outcome in iter_outcomes((data.get("resourceOutcomes", {}) or {})):
        for msg in (outcome.get("messages", []) or []):
            if not isinstance(msg, dict):
                continue
            ws.append([
                f"msg_{msg_index:04d}",
                outcome_id,
                msg.get("channel", ""),
                msg.get("actor", ""),
                msg.get("delay", ""),
                msg.get("type", "text"),
                msg.get("content", ""),
                msg.get("hint", ""),
                msg.get("outcome", ""),
            ])
            msg_index += 1
    finalize_sheet(ws)
    
    # EVENTS_CINEMATICS
    ws = ensure_sheet(wb, "EVENTS_CINEMATICS", [
        "event_id", "case_id", "trigger_condition", "event_type", "title",
        "body", "priority", "media_src"
    ])
    for event in (data.get("cinematicEvents", []) or []):
        ws.append([
            event.get("id", ""),
            event.get("case", ""),
            "on_case_start",
            "cinematic",
            event.get("title", ""),
            event.get("body", ""),
            "1",
            event.get("media", ""),
        ])
    finalize_sheet(ws)
    
    # ENDINGS
    ws = ensure_sheet(wb, "ENDINGS", [
        "ending_id", "condition_expr", "ending_type", "title", "body",
        "score_delta", "next_game"
    ])
    for ending_id, ending_data in (data.get("endings", {}) or {}).items():
        ws.append([
            ending_id,
            "",
            "victory",
            ending_data.get("title", ""),
            ending_data.get("body", ""),
            "0",
            "",
        ])
    finalize_sheet(ws)
    
    # ASSETS
    ws = ensure_sheet(wb, "ASSETS", [
        "asset_id", "asset_type", "path", "license", "fallback_asset_id", "description"
    ])
    for asset_id, asset_data in ((data.get("assets", {}) or {}).get("music", {}) or {}).items():
        ws.append([asset_id, "music", asset_data.get("path", ""), "", "", ""])
    finalize_sheet(ws)
    
    # VALIDATION
    ws = ensure_sheet(wb, "VALIDATION", ["Rule Type", "Scope", "Check", "Error Message"])
    finalize_sheet(ws)
    
    # Reorder sheets
    wb._sheets = [wb[s] for s in SHEETS_ORDER_V2 if s in wb.sheetnames]
    apply_editor_guided_validations(wb)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)

# ============================================================================
# VALIDATION ENGINE
# ============================================================================

class ValidationError(Exception):
    """Validation error with context."""
    def __init__(self, sheet: str, row: int, column: str, message: str):
        self.sheet = sheet
        self.row = row
        self.column = column
        self.message = message
    
    def __str__(self):
        return f"[{self.sheet}:{self.row}:{self.column}] {self.message}"

class SchemaValidator:
    """Validate game schema for referential integrity and correctness."""
    
    def __init__(self):
        self.errors: List[ValidationError] = []
        self.warnings: List[str] = []
    
    def validate_schema(self, data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Run full schema validation. Returns (is_valid, error_messages)."""
        self.errors = []
        self.warnings = []
        
        self._validate_characters(data)
        self._validate_channels(data)
        self._validate_cases(data)
        self._validate_references(data)
        
        error_messages = [str(e) for e in self.errors]
        return len(self.errors) == 0, error_messages + self.warnings
    
    def _validate_characters(self, data: Dict[str, Any]) -> None:
        """Validate character definitions."""
        characters = data.get("characters", {})
        char_ids = set(characters.keys())
        
        if not char_ids:
            self.warnings.append("⚠ No characters defined. Game may not run.")
            return
        
        for char_id in char_ids:
            if not char_id or not isinstance(char_id, str):
                self.errors.append(ValidationError("CHARACTERS", 0, "character_id", "Character ID must be non-empty string"))
            if " " in char_id:
                self.errors.append(ValidationError("CHARACTERS", 0, "character_id", f"Character ID '{char_id}' contains spaces"))
    
    def _validate_channels(self, data: Dict[str, Any]) -> None:
        """Validate channel definitions."""
        channels = data.get("channels", [])
        channel_ids = set()
        
        for channel_entry in channels:
            if not isinstance(channel_entry, (list, tuple)) or len(channel_entry) < 2:
                continue
            channel_id = channel_entry[0]
            channel_data = channel_entry[1]
            
            if not channel_id or not isinstance(channel_id, str):
                self.errors.append(ValidationError("CHANNELS", 0, "channel_id", "Channel ID must be non-empty string"))
                continue
            
            channel_ids.add(channel_id)
            
            if " " in channel_id:
                self.errors.append(ValidationError("CHANNELS", 0, "channel_id", f"Channel ID '{channel_id}' contains spaces"))
            
            # Validate actor reference
            actor = channel_data.get("actor", "")
            if actor and actor not in data.get("characters", {}):
                self.errors.append(ValidationError("CHANNELS", 0, "actor_id", 
                    f"Channel '{channel_id}' references unknown character '{actor}'"))
        
        if not channel_ids:
            self.warnings.append("⚠ No channels defined. Game may not run.")
    
    def _validate_cases(self, data: Dict[str, Any]) -> None:
        """Validate case definitions."""
        cases = data.get("cases", [])
        case_ids = set()
        characters = set(data.get("characters", {}).keys())
        
        for case_idx, case in enumerate(cases, start=1):
            if not isinstance(case, dict):
                continue
            
            case_id = case.get("id", "")
            if not case_id or not isinstance(case_id, str):
                self.errors.append(ValidationError("CASES", case_idx, "case_id", "Case ID must be non-empty string"))
                continue
            
            case_ids.add(case_id)
            
            if " " in case_id:
                self.errors.append(ValidationError("CASES", case_idx, "case_id", f"Case ID '{case_id}' contains spaces"))

            status = str(case.get("status", "pending") or "pending").strip().lower()
            if status not in {"pending", "resolved"}:
                self.errors.append(ValidationError("CASES", case_idx, "status",
                    f"Case '{case_id}' has invalid status '{status}' (expected pending/resolved)"))

            schedule_time = str(case.get("scheduleTime", "") or "").strip()
            if schedule_time and not re.match(r"^([01]?\d|2[0-3]):[0-5]\d$", schedule_time):
                self.errors.append(ValidationError("CASES", case_idx, "schedule_time",
                    f"Case '{case_id}' has invalid schedule_time '{schedule_time}' (expected HH:MM)"))
            
            # Validate conversations
            conversations = case.get("conversations", [])
            linked_characters = set(case.get("characterIds", []) or [])
            if not linked_characters and isinstance(case.get("caseCharacters"), list):
                linked_characters = {
                    str(item.get("character_id") or item.get("characterId") or "").strip()
                    for item in (case.get("caseCharacters") or []) if isinstance(item, dict)
                }
            for conv_idx, conv in enumerate(conversations, start=1):
                if not isinstance(conv, dict):
                    continue
                
                channel_id = conv.get("channel", "")
                if channel_id and not any(ch[0] == channel_id for ch in data.get("channels", [])):
                    self.errors.append(ValidationError("CONVERSATIONS", conv_idx, "channel_id",
                        f"Case '{case_id}' references unknown channel '{channel_id}'"))
                
                actor_id = conv.get("actor", "")
                if actor_id and actor_id not in characters and actor_id not in ("mimi_protagonist", "system"):
                    self.warnings.append(f"⚠ Conversation in case '{case_id}' references unknown actor '{actor_id}'")
                if linked_characters and actor_id and actor_id not in linked_characters and actor_id not in ("mimi_protagonist", "system"):
                    self.errors.append(ValidationError("CONVERSATIONS", conv_idx, "actor_id",
                        f"Case '{case_id}' has actor '{actor_id}' not linked in CASE_CHARACTERS"))

            if not linked_characters:
                self.warnings.append(f"⚠ Case '{case_id}' has no explicit CASE_CHARACTERS links")
            
            # Validate signal references
            required_signals = case.get("requiredSignals", [])
            signal_responses = case.get("signalResponse", {})
            
            for signal in required_signals:
                if signal not in signal_responses:
                    self.errors.append(ValidationError("SIGNALS_INFERENCE", case_idx, "signal_key",
                        f"Case '{case_id}': required signal '{signal}' has no reveal_text in signalResponse"))

            # Validate authored dialogues against explicit links
            for dlg_idx, dialogue in enumerate((case.get("dialogues", []) or []), start=1):
                if not isinstance(dialogue, dict):
                    continue
                actor_id = str(dialogue.get("character_id") or dialogue.get("characterId") or dialogue.get("actor") or "").strip()
                if actor_id and linked_characters and actor_id not in linked_characters:
                    self.errors.append(ValidationError("DIALOGUES", dlg_idx, "character_id",
                        f"Case '{case_id}' dialogue actor '{actor_id}' is not linked in CASE_CHARACTERS"))
    
    def _validate_references(self, data: Dict[str, Any]) -> None:
        """Validate cross-sheet references."""
        case_ids = {c.get("id") for c in data.get("cases", []) if c.get("id")}
        asset_ids = set()
        
        # Asset references
        for asset_id in data.get("assets", {}).get("music", {}).keys():
            asset_ids.add(asset_id)
        
        # Resource outcome references
        for outcome_id, outcome in data.get("resourceOutcomes", {}).items():
            if not isinstance(outcome, dict):
                continue
            
            resource_id = outcome.get("resource", "")
            if resource_id and resource_id not in data.get("resources", {}):
                self.warnings.append(f"⚠ Outcome '{outcome_id}' references unknown resource '{resource_id}'")
            
            branch = outcome.get("branch", "")
            if branch and branch not in {"success", "partial", "failure"}:
                self.errors.append(ValidationError("RESOURCE_OUTCOMES", 0, "branch",
                    f"Outcome '{outcome_id}': branch '{branch}' must be success/partial/failure"))

# ============================================================================
# IMPORT FROM WORKBOOK
# ============================================================================

def rows_as_dicts(ws: Worksheet, skip_comments: bool = True) -> List[Dict[str, Any]]:
    """Convert worksheet to list of dicts. Optionally skip comment/example rows."""
    if ws is None:
        return []
    
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    out: List[Dict[str, Any]] = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        
        # Skip comment rows (first col contains long descriptive text)
        if skip_comments:
            first_val = row[0]
            if first_val and len(str(first_val)) > 100:  # Long text = comment row
                continue
            # Skip example rows (marked by colored cells in export)
            if first_val and ("ejemplo" in str(first_val).lower() or "example" in str(first_val).lower()):
                continue
        
        out.append({headers[i]: row[i] for i in range(len(headers))})
    
    return out

def row_value(row: Dict[str, Any], aliases: Iterable[str], default: Any = "") -> Any:
    """Get first non-empty value from a row by trying multiple column aliases."""
    for key in aliases:
        if key not in row:
            continue
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return default

def import_from_workbook_v2(wb_path: Path, json_base: Dict[str, Any], validate: bool = True) -> Tuple[Dict[str, Any], List[str]]:
    """Import v2 Excel to game JSON. Returns (data, validation_messages)."""
    wb = load_workbook(wb_path)
    out = copy.deepcopy(json_base)
    validation_messages: List[str] = []
    
    def get_sheet(name: str) -> Optional[Worksheet]:
        for candidate in SHEET_NAME_ALIASES.get(name, [name]):
            if candidate in wb.sheetnames:
                return wb[candidate]
        return None

    # CORE_META
    meta_rows = rows_as_dicts(get_sheet("CORE_META") or wb.active)
    meta: Dict[str, Any] = {}
    for row in meta_rows:
        key = str(row.get("key", row.get("field", ""))).strip()
        if not key:
            continue
        meta[key] = row.get("value")

    game_domain = str(meta.get("game_domain", "")).strip()
    if game_domain:
        out["game_domain"] = game_domain
    game_title = str(meta.get("game_title", "")).strip()
    if game_title:
        out["game_title"] = game_title
    game_version = str(meta.get("game_version", meta.get("version", ""))).strip()
    if game_version:
        out["version"] = game_version

    current_strict = as_bool((out.get("settings", {}) or {}).get("interaction", {}).get("strictAuthoring"), True)
    out.setdefault("settings", {}).setdefault("interaction", {})["strictAuthoring"] = as_bool(meta.get("strict_authoring", meta.get("strictAuthoring")), current_strict)
    player_label = str(meta.get("player_label", meta.get("playerLabel", ""))).strip()
    if player_label:
        out.setdefault("settings", {}).setdefault("ui", {})["playerLabel"] = player_label
    archive_label = str(meta.get("archive_label", meta.get("archiveLabel", ""))).strip()
    if archive_label:
        out.setdefault("settings", {}).setdefault("ui", {})["archiveLabel"] = archive_label

    no_match_responses = str(meta.get("no_match_responses", meta.get("noMatchResponses", ""))).strip()
    if no_match_responses:
        out.setdefault("settings", {}).setdefault("semantic", {})["noMatchResponses"] = csv_split(no_match_responses)

    interaction = out.setdefault("settings", {}).setdefault("interaction", {})
    turn_start_time = str(meta.get("turn_start_time", meta.get("turnStartTime", meta.get("start_time", ""))) or "").strip()
    if turn_start_time:
        interaction["turnStartTime"] = turn_start_time
    turn_end_time = str(meta.get("turn_end_time", meta.get("turnEndTime", meta.get("end_time", ""))) or "").strip()
    if turn_end_time:
        interaction["turnEndTime"] = turn_end_time
    tick_ms = row_value(meta, ["clock_tick_interval_ms", "clockTickIntervalMs"], None)
    if tick_ms is not None and str(tick_ms).strip() != "":
        interaction["clockTickIntervalMs"] = as_int(tick_ms, 7000)
    slow_minutes = row_value(meta, ["clock_advance_slow_minutes", "clockAdvanceSlowMinutes"], None)
    if slow_minutes is not None and str(slow_minutes).strip() != "":
        interaction["clockAdvanceSlowMinutes"] = as_int(slow_minutes, 4)
    normal_minutes = row_value(meta, ["clock_advance_normal_minutes", "clockAdvanceNormalMinutes"], None)
    if normal_minutes is not None and str(normal_minutes).strip() != "":
        interaction["clockAdvanceNormalMinutes"] = as_int(normal_minutes, 6)
    rush_minutes = row_value(meta, ["clock_advance_rush_minutes", "clockAdvanceRushMinutes"], None)
    if rush_minutes is not None and str(rush_minutes).strip() != "":
        interaction["clockAdvanceRushMinutes"] = as_int(rush_minutes, 9)
    
    # CHARACTERS
    char_rows = rows_as_dicts(get_sheet("CHARACTERS") or wb.active)
    characters: Dict[str, Any] = {}
    for row in char_rows:
        char_id = str(row_value(row, CHARACTER_HEADER_ALIASES["character_id"], "")).strip()
        if not char_id:
            continue
        characters[char_id] = {
            "name": str(row_value(row, CHARACTER_HEADER_ALIASES["name"], "") or ""),
            "role": str(row_value(row, CHARACTER_HEADER_ALIASES["role"], "") or ""),
            "voice": str(row_value(row, CHARACTER_HEADER_ALIASES["voice"], "") or ""),
            "description": str(row_value(row, CHARACTER_HEADER_ALIASES["description"], "") or ""),
        }
    out["characters"] = characters
    protagonist_entry = next(((char_id, value) for char_id, value in characters.items() if str(value.get("role", "")).strip().lower() == "protagonist"), None)
    if protagonist_entry:
        protagonist_id, protagonist_data = protagonist_entry
        out["protagonist"] = {
            "id": protagonist_id,
            "name": protagonist_data.get("name") or protagonist_id.replace("_", " ").title(),
            "phoneChats": list((out.get("protagonist", {}) or {}).get("phoneChats", [])),
        }
    
    # CHANNELS
    channel_rows = rows_as_dicts(get_sheet("CHANNELS") or wb.active)
    channels: List[Tuple[str, Dict[str, Any]]] = []
    for row in channel_rows:
        channel_id = str(row_value(row, CHANNEL_HEADER_ALIASES["channel_id"], "")).strip()
        if not channel_id:
            continue
        channels.append((channel_id, {
            "type": str(row_value(row, CHANNEL_HEADER_ALIASES["type"], "") or ""),
            "actor": str(row_value(row, CHANNEL_HEADER_ALIASES["actor"], "") or ""),
            "label": str(row_value(row, CHANNEL_HEADER_ALIASES["label"], "") or ""),
            "color": str(row_value(row, CHANNEL_HEADER_ALIASES["color"], "") or ""),
            "description": str(row_value(row, CHANNEL_HEADER_ALIASES["description"], "") or ""),
        }))
    out["channels"] = channels
    
    # CASES
    case_rows = rows_as_dicts(get_sheet("CASES") or wb.active)
    cases: List[Dict[str, Any]] = []
    for row in case_rows:
        case_id = str(row_value(row, CASE_HEADER_ALIASES["case_id"], "")).strip()
        if not case_id:
            continue
        opening_line = row_value(row, CASE_HEADER_ALIASES["opening_line"], "")
        opening_text = str(opening_line or "").strip()
        cases.append({
            "id": case_id,
            "title": str(row_value(row, CASE_HEADER_ALIASES["title"], "") or ""),
            "status": str(row_value(row, CASE_HEADER_ALIASES["status"], "pending") or "pending").strip().lower(),
            "category": str(row_value(row, CASE_HEADER_ALIASES["category"], "") or ""),
            "scheduleTime": str(row_value(row, CASE_HEADER_ALIASES["schedule_time"], "") or ""),
            "allocationBucket": str(row_value(row, CASE_HEADER_ALIASES["allocation_bucket"], "global") or "global").strip(),
            "minimumClues": as_int(row_value(row, CASE_HEADER_ALIASES["minimum_clues"], 1), 1),
            "severity": str(row_value(row, CASE_HEADER_ALIASES["severity"], "media") or "media"),
            "stress": as_int(row_value(row, CASE_HEADER_ALIASES["stress"], 0), 0),
            "openingLine": opening_text,
            "narrativeIntro": csv_split(row_value(row, CASE_HEADER_ALIASES["narrative_intro"], "")) or ([opening_text] if opening_text else []),
            "optimal": delimited_split(row_value(row, CASE_HEADER_ALIASES["optimal_resources"], "")),
            "conversations": [],
            "signalResponse": {},
            "requiredSignals": [],
            "characterIds": [],
            "caseCharacters": [],
            "dialogues": [],
        })
    out["cases"] = cases

    # CASE_CHARACTERS
    case_char_rows = rows_as_dicts(get_sheet("CASE_CHARACTERS") or wb.active)
    case_char_by_case: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    case_ids_by_case: Dict[str, List[str]] = defaultdict(list)
    for row in case_char_rows:
        case_id = str(row.get("case_id", "")).strip()
        character_id = str(row.get("character_id", "")).strip()
        if not case_id or not character_id:
            continue
        case_char_by_case[case_id].append({
            "character_id": character_id,
            "involvement_role": str(row.get("involvement_role", "participant") or "participant"),
            "is_primary": as_bool(row.get("is_primary"), False),
        })
        if character_id not in case_ids_by_case[case_id]:
            case_ids_by_case[case_id].append(character_id)
    
    # CONVERSATIONS
    conv_rows = rows_as_dicts(get_sheet("CONVERSATIONS") or wb.active)
    conv_by_case: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in conv_rows:
        case_id = str(row_value(row, CONVERSATION_HEADER_ALIASES["case_id"], "")).strip()
        if not case_id:
            continue
        conv_by_case[case_id].append({
            "channel": str(row_value(row, CONVERSATION_HEADER_ALIASES["channel"], "") or ""),
            "actor": str(row_value(row, CONVERSATION_HEADER_ALIASES["actor"], "") or ""),
            "time": str(row_value(row, CONVERSATION_HEADER_ALIASES["time"], "") or ""),
            "type": str(row_value(row, CONVERSATION_HEADER_ALIASES["type"], "text") or "text"),
            "content": str(row_value(row, CONVERSATION_HEADER_ALIASES["content"], "") or ""),
            "src": str(row_value(row, CONVERSATION_HEADER_ALIASES["src"], "") or ""),
            "alt": str(row_value(row, CONVERSATION_HEADER_ALIASES["alt"], "") or ""),
        })
    
    for case in out.get("cases", []):
        case_id = case.get("id")
        case_conversations = conv_by_case.get(case_id, [])
        case["conversations"] = case_conversations
        explicit_links = case_char_by_case.get(case_id, [])
        explicit_ids = case_ids_by_case.get(case_id, [])
        case["caseCharacters"] = explicit_links
        case["characterIds"] = explicit_ids

    # DIALOGUES
    dialogue_rows = rows_as_dicts(get_sheet("DIALOGUES") or wb.active)
    dialogues_by_case: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    reveal_from_dialogues: Dict[str, Dict[str, str]] = defaultdict(dict)
    for row in dialogue_rows:
        case_id = str(row.get("case_id", "")).strip()
        if not case_id:
            continue
        keyword = str(row.get("keyword", "")).strip()
        response = str(row.get("response", "") or "")
        clue = str(row.get("clue", "") or "")
        item = {
            "character_id": str(row.get("character_id", "") or ""),
            "intent": str(row.get("intent", "") or ""),
            "keyword": keyword,
            "response": response,
            "clue": clue,
            "resource_id": str(row.get("resource_id", "") or ""),
        }
        dialogues_by_case[case_id].append(item)
        if keyword and response:
            reveal_from_dialogues[case_id][keyword] = response

    for case in out.get("cases", []):
        case_id = case.get("id")
        case["dialogues"] = dialogues_by_case.get(case_id, [])
        if reveal_from_dialogues.get(case_id):
            case["signalResponse"] = reveal_from_dialogues[case_id]
            if not case.get("requiredSignals"):
                case["requiredSignals"] = list(reveal_from_dialogues[case_id].keys())
    
    # SIGNALS_INFERENCE
    signal_rows = rows_as_dicts(get_sheet("SIGNALS_INFERENCE") or wb.active)
    signal_by_case: Dict[str, Dict[str, str]] = defaultdict(dict)
    critical_by_case: Dict[str, List[str]] = defaultdict(list)
    for row in signal_rows:
        case_id = str(row_value(row, SIGNAL_HEADER_ALIASES["case_id"], "")).strip()
        signal_key = str(row_value(row, SIGNAL_HEADER_ALIASES["signal_key"], "")).strip()
        if not case_id or not signal_key:
            continue
        reveal_text = str(row_value(row, SIGNAL_HEADER_ALIASES["reveal_text"], "") or "")
        signal_by_case[case_id][signal_key] = reveal_text
        
        if as_bool(row_value(row, SIGNAL_HEADER_ALIASES["is_critical"], False), False):
            critical_by_case[case_id].append(signal_key)
    
    for case in out.get("cases", []):
        case_id = case.get("id", "")
        case["signalResponse"] = signal_by_case.get(case_id, {})
        case["requiredSignals"] = critical_by_case.get(case_id, [])
    
    # INTENTS_ALIASES
    alias_rows = rows_as_dicts(get_sheet("INTENTS_ALIASES") or wb.active)
    aliases: Dict[str, List[str]] = {}
    for row in alias_rows:
        canonical = str(row.get("canonical_intent", row.get("signal_key", ""))).strip()
        if not canonical:
            continue
        if "aliases" in row:
            alias_values = csv_split(row.get("aliases", ""))
        else:
            alias_values = [str(row.get("alias", "") or "").strip()] if str(row.get("alias", "") or "").strip() else []
        aliases.setdefault(canonical, [])
        for alias in alias_values:
            if alias and alias not in aliases[canonical]:
                aliases[canonical].append(alias)
    
    out.setdefault("settings", {}).setdefault("semantic", {})["keywordAliases"] = aliases

    # RESOURCES
    resource_rows = rows_as_dicts(get_sheet("RESOURCES") or wb.active)
    resources: Dict[str, Any] = {}
    for row in resource_rows:
        resource_id = str(row.get("resource_id", "")).strip()
        if not resource_id:
            continue
        resources[resource_id] = {
            "label": str(row.get("label", "") or ""),
            "description": str(row.get("description", "") or ""),
            "category": str(row.get("category", "tool") or "tool"),
            "total": as_int(row.get("total_available") if row.get("total_available") is not None else row.get("total"), 0),
            "cost": as_int(row.get("cost"), 1),
            "allocationMode": str((row.get("allocation_mode") if row.get("allocation_mode") is not None else row.get("allocationMode")) or "global"),
            "capacities": {},
        }

    capacity_rows = rows_as_dicts(get_sheet("RESOURCE_CAPACITY") or wb.active)
    for row in capacity_rows:
        resource_id = str(row.get("resource_id", "") or "").strip()
        bucket_key = str((row.get("bucket_key") if row.get("bucket_key") is not None else row.get("bucket")) or "").strip()
        if not resource_id or not bucket_key or resource_id not in resources:
            continue
        amount = row.get("available") if row.get("available") is not None else row.get("capacity")
        resources[resource_id].setdefault("capacities", {})[bucket_key] = as_int(amount, 0)
    if resources:
        out["resources"] = resources
    
    # ASSETS
    asset_rows = rows_as_dicts(get_sheet("ASSETS") or wb.active)
    music_assets: Dict[str, Any] = {}
    for row in asset_rows:
        asset_id = str(row.get("asset_id", "")).strip()
        if not asset_id:
            continue
        asset_type = str(row.get("asset_type", "") or "").lower()
        if asset_type == "music":
            music_assets[asset_id] = {
                "path": str(row.get("path", "") or ""),
                "license": str(row.get("license", "") or ""),
            }
    
    out.setdefault("assets", {})["music"] = music_assets

    # PHONE_CONTACTS / PHONE_MESSAGES
    phone_contact_rows = rows_as_dicts(get_sheet("PHONE_CONTACTS") or wb.active)
    phone_chats: Dict[str, Dict[str, Any]] = {}
    for row in phone_contact_rows:
        contact_id = str(row.get("contact_id", "") or "").strip()
        if not contact_id:
            continue
        media_permissions = delimited_split(
            row.get("media_permissions")
            if row.get("media_permissions") is not None
            else row.get("mediaPermissions", "")
        )
        phone_chats[contact_id] = {
            "id": contact_id,
            "contact": str(row.get("contact", "") or contact_id),
            "avatar": str(row.get("avatar", "") or ""),
            "color": str(row.get("color", "") or ""),
            "status": str(row.get("status", "") or ""),
            "unread": as_int(row.get("unread"), 0),
            "autoReply": str(row.get("autoReply", "") or ""),
            "mediaPermissions": media_permissions,
            "messages": [],
        }

    phone_message_rows = rows_as_dicts(get_sheet("PHONE_MESSAGES") or wb.active)
    for row in phone_message_rows:
        contact_id = str(row.get("contact_id", "") or "").strip()
        if not contact_id:
            continue
        if contact_id not in phone_chats:
            phone_chats[contact_id] = {
                "id": contact_id,
                "contact": contact_id,
                "avatar": "",
                "color": "",
                "status": "",
                "unread": 0,
                "autoReply": "",
                "messages": [],
            }
        phone_chats[contact_id]["messages"].append({
            "from": str(row.get("from", "them") or "them"),
            "time": str(row.get("time", "") or ""),
            "text": str(row.get("text", "") or ""),
            "type": str(row.get("type", "text") or "text"),
            "src": str(row.get("src", "") or ""),
            "tone": str(row.get("tone", "") or ""),
        })

    if phone_chats:
        protagonist = out.setdefault("protagonist", {})
        protagonist.setdefault("name", out.get("settings", {}).get("ui", {}).get("playerLabel") or "Operador")
        protagonist["phoneChats"] = list(phone_chats.values())
    
    # RESOURCE_OUTCOMES
    outcome_rows = rows_as_dicts(get_sheet("RESOURCE_OUTCOMES") or wb.active)
    outcomes: Dict[str, Any] = {}
    for row in outcome_rows:
        outcome_id = str(row.get("outcome_id", "")).strip()
        if not outcome_id:
            continue
        branch = str(row.get("branch", "") or "").strip().lower()
        if not branch:
            if as_bool(row.get("success"), False):
                branch = "success"
            elif as_bool(row.get("partial"), False):
                branch = "partial"
            else:
                branch = "failure"
        inline_message = str(row.get("message_content", "") or "")
        clue_notes = csv_split(row.get("clueNotes", ""))
        clues_unlocked = csv_split(row.get("cluesUnlocked", ""))
        outcomes[outcome_id] = {
            "case_id": str(row.get("case_id", "") or ""),
            "resource": str(row.get("resource_id", "") or ""),
            "branch": branch,
            "confidence_band": str(row.get("confidence_band", "") or ""),
            "unlocks": csv_split(row.get("unlocks_signals", row.get("cluesUnlocked", ""))),
            "score_delta": as_int(row.get("score_delta") if row.get("score_delta") is not None else row.get("scoreDelta"), 0),
            "calm_delta": as_int(row.get("calm_delta") if row.get("calm_delta") is not None else row.get("calmDelta"), 0),
            "next_event": str((row.get("next_event") if row.get("next_event") is not None else row.get("nextEvent")) or ""),
            "keep_case_open": as_bool(row.get("keep_case_open") if row.get("keep_case_open") is not None else row.get("keepCaseOpen"), False),
            "reaction": str(row.get("reaction", "") or ""),
            "clueNotes": clue_notes,
            "cluesUnlocked": clues_unlocked,
            "messages": [],
        }
        if inline_message:
            outcomes[outcome_id]["messages"].append({
                "channel": "",
                "actor": "system",
                "delay": 0,
                "type": "text",
                "content": inline_message,
                "hint": clues_unlocked[0] if clues_unlocked else "",
                "outcome": "",
            })

    # OUTCOME_MESSAGES
    outcome_msg_rows = rows_as_dicts(get_sheet("OUTCOME_MESSAGES") or wb.active)
    for row in outcome_msg_rows:
        outcome_id = str(row.get("outcome_id", "")).strip()
        if not outcome_id or outcome_id not in outcomes:
            continue
        outcomes[outcome_id]["messages"].append({
            "channel": str((row.get("channel_id") if row.get("channel_id") is not None else row.get("channel")) or ""),
            "actor": str((row.get("actor_id") if row.get("actor_id") is not None else row.get("actor")) or ""),
            "delay": as_int(row.get("delay_ms") if row.get("delay_ms") is not None else row.get("delay"), 0),
            "type": str((row.get("message_type") if row.get("message_type") is not None else row.get("type")) or "text"),
            "content": str(row.get("content", "") or ""),
            "hint": str((row.get("hint_key") if row.get("hint_key") is not None else row.get("hint")) or ""),
            "outcome": str((row.get("next_case") if row.get("next_case") is not None else row.get("outcome")) or ""),
        })

    out["resourceOutcomes"] = outcomes
    
    # CINEMATICS
    cinematic_rows = rows_as_dicts(get_sheet("EVENTS_CINEMATICS") or wb.active)
    cinematics: List[Dict[str, Any]] = []
    for row in cinematic_rows:
        event_id = str(row.get("event_id", row.get("id", ""))).strip()
        if not event_id:
            continue
        cinematics.append({
            "id": event_id,
            "case": str(row.get("case_id", row.get("triggerCase", "")) or ""),
            "title": str(row.get("title", row.get("description", "")) or ""),
            "body": str(row.get("body", row.get("effect", "")) or ""),
            "media": str(row.get("media_src", "") or ""),
        })
    out.setdefault("cinematicEvents", cinematics)
    
    # ENDINGS
    ending_rows = rows_as_dicts(get_sheet("ENDINGS") or wb.active)
    endings: Dict[str, Dict[str, Any]] = {}
    for row in ending_rows:
        ending_id = str(row.get("ending_id", "")).strip()
        if not ending_id:
            continue
        endings[ending_id] = {
            "title": str(row.get("title", "") or ""),
            "body": str(row.get("body", "") or ""),
        }
    out.setdefault("endings", endings)
    
    # SCHEMA
    schema = out.setdefault("settings", {}).setdefault("schema", {})
    schema["version"] = SCHEMA_VERSION
    schema["compatRuntimeVersion"] = COMPAT_RUNTIME_VERSION
    schema["cores"] = RUNTIME_CORES
    
    # VALIDATION
    if validate:
        validator = SchemaValidator()
        is_valid, messages = validator.validate_schema(out)
        validation_messages.extend(messages)
        if not is_valid:
            print(f"⛔ Validation failed with {len([m for m in messages if m.startswith('[')])} errors:")
            for msg in messages:
                if msg.startswith('['):
                    print(f"  {msg}")
    
    return out, validation_messages

# ============================================================================
# TEMPLATE GENERATION
# ============================================================================

def create_template_only(xlsx_path: Path) -> None:
    """Create blank v2 template without data."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # HOME sheet (no headers, just content)
    ws = wb.create_sheet(title="HOME", index=0)
    write_home(ws)
    
    ws = ensure_sheet(wb, "CORE_META", ["key", "value", "description"])
    write_core_meta(ws)
    
    ws = ensure_sheet(wb, "CHARACTERS", ["character_id", "display_name", "role", "voice_style", "description"])
    write_characters(ws)
    
    ws = ensure_sheet(wb, "CHANNELS", ["channel_id", "channel_type", "actor_id", "ui_label", "ui_color", "description"])
    write_channels(ws)
    
    ws = ensure_sheet(wb, "CASES", [
        "case_id", "title", "status", "category", "schedule_time", "allocation_bucket",
        "minimum_clues", "severity", "stress", "opening_line"
    ])
    write_cases(ws)

    ws = ensure_sheet(wb, "CASE_CHARACTERS", ["case_id", "character_id", "involvement_role", "is_primary"])
    write_case_characters(ws)

    ws = ensure_sheet(wb, "DIALOGUES", ["case_id", "character_id", "intent", "keyword", "response", "clue", "resource_id"])
    write_dialogues(ws)
    
    ws = ensure_sheet(wb, "CONVERSATIONS", [
        "conversation_id", "case_id", "channel_id", "order_index", "actor_id",
        "time_hhmm", "message_type", "content", "media_src", "media_alt", "visible_if_result"
    ])
    write_conversations(ws)
    
    ws = ensure_sheet(wb, "INTENTS_ALIASES", ["canonical_intent", "display_label", "aliases", "description"])
    write_intents_aliases(ws)
    
    ws = ensure_sheet(wb, "SIGNALS_INFERENCE", [
        "signal_key", "case_id", "reveal_text", "is_critical"
    ])
    write_signals_inference(ws)
    
    ws = ensure_sheet(wb, "RESOURCES", [
        "resource_id", "label", "description", "category", "total_available", "cost", "allocation_mode"
    ])
    write_resources(ws)

    ws = ensure_sheet(wb, "RESOURCE_CAPACITY", ["resource_id", "bucket_key", "available", "travel_ms", "description"])
    write_resource_capacity(ws)
    
    ws = ensure_sheet(wb, "RESOURCE_OUTCOMES", [
        "outcome_id", "case_id", "resource_id", "branch", "confidence_band",
        "unlocks_signals", "score_delta", "calm_delta", "next_event", "keep_case_open"
    ])
    write_resource_outcomes(ws)
    
    ws = ensure_sheet(wb, "OUTCOME_MESSAGES", [
        "message_id", "outcome_id", "channel_id", "actor_id", "delay_ms",
        "message_type", "content", "hint_key", "next_case"
    ])
    write_outcome_messages(ws)
    
    ws = ensure_sheet(wb, "EVENTS_CINEMATICS", [
        "event_id", "case_id", "trigger_condition", "event_type", "title",
        "body", "priority", "media_src"
    ])
    write_events_cinematics(ws)
    
    ws = ensure_sheet(wb, "ENDINGS", [
        "ending_id", "condition_expr", "ending_type", "title", "body",
        "score_delta", "next_game"
    ])
    write_endings(ws)
    
    ws = ensure_sheet(wb, "ASSETS", [
        "asset_id", "asset_type", "path", "license", "fallback_asset_id", "description"
    ])
    write_assets(ws)
    
    ws = ensure_sheet(wb, "VALIDATION", ["Rule Type", "Scope", "Check", "Error Message"])
    write_validation(ws)
    
    wb._sheets = [wb[s] for s in SHEETS_ORDER_V2 if s in wb.sheetnames]
    apply_editor_guided_validations(wb)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"Template created: {xlsx_path}")

# ============================================================================
# CLI
# ============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Domain-agnostic Excel ↔ JSON bridge v2.")
    sub = parser.add_subparsers(dest="command", required=True)
    
    ex = sub.add_parser("export", help="Export JSON to Excel v2.")
    ex.add_argument("--json", required=True, type=Path, help="Input JSON file.")
    ex.add_argument("--xlsx", required=True, type=Path, help="Output XLSX file.")
    
    im = sub.add_parser("import", help="Import Excel v2 to JSON.")
    im.add_argument("--xlsx", required=True, type=Path, help="Input XLSX file.")
    im.add_argument("--json-out", required=True, type=Path, help="Output JSON file.")
    im.add_argument("--json-base", required=True, type=Path, help="Base JSON to preserve unrelated fields.")
    
    tm = sub.add_parser("template", help="Create blank v2 template.")
    tm.add_argument("--xlsx", required=True, type=Path, help="Output template file.")
    
    return parser.parse_args()

def main() -> None:
    args = parse_args()
    
    if args.command == "export":
        data = load_json(args.json)
        validator = SchemaValidator()
        is_valid, messages = validator.validate_schema(data)
        if not is_valid:
            print(f"⛔ Source JSON has {len([m for m in messages if m.startswith('[')])} validation errors. Fix before export:")
            for msg in messages:
                if msg.startswith('['):
                    print(f"  {msg}")
            return
        export_to_workbook_v2(data, args.xlsx)
        print(f"✓ Exported to: {args.xlsx}")
        if [m for m in messages if m.startswith('⚠')]:
            for msg in messages:
                if msg.startswith('⚠'):
                    print(f"  {msg}")
    
    elif args.command == "import":
        base = load_json(args.json_base)
        output, validation_messages = import_from_workbook_v2(args.xlsx, base, validate=True)
        save_json(args.json_out, output)
        print(f"✓ Imported to: {args.json_out}")
        
        errors = [m for m in validation_messages if m.startswith('[')]
        warnings = [m for m in validation_messages if m.startswith('⚠')]
        
        if errors:
            print(f"\n⛔ {len(errors)} validation error(s):")
            for msg in errors:
                print(f"  {msg}")
        
        if warnings:
            print(f"\n⚠ {len(warnings)} warning(s):")
            for msg in warnings:
                print(f"  {msg}")
    
    elif args.command == "template":
        create_template_only(args.xlsx)

if __name__ == "__main__":
    main()
